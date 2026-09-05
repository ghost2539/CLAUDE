"""Shared helper functions used across multiple routers."""
from __future__ import annotations

import io
import json
from datetime import date
from decimal import Decimal
from typing import Any

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from sqlalchemy import select, func, or_
from sqlalchemy.orm import Session

from database import (
    Asset, Classification, LocalAsset, ReceiptCycle, StorageLocation,
)


# ── Classification ────────────────────────────────────────────────

def classify(
    s: Session,
    description: str,
    company: str = "",
) -> tuple[str, str, str]:
    """Return (category, model, company) for a given description."""
    q = (
        select(Classification)
        .where(Classification.active == True)  # noqa: E712
        .order_by(func.length(Classification.description_pattern).desc())
    )
    rules = s.scalars(q).all()
    d = (description or "").upper()
    c = (company or "").upper()
    for r in rules:
        if r.description_pattern.upper() in d and (not r.company or r.company.upper() == c):
            return r.category, r.model, r.company or company
    return "NÃO CLASSIFICADA", description or "", company


def apply_class(s: Session, r: dict) -> dict:
    """Apply classification rules to a search result dict."""
    if r.get("encontrado"):
        cat, model, company = classify(s, r.get("descricao", ""), r.get("empresa", ""))
        r["categoria"] = cat
        r["modelo"] = model
        r["empresa"] = company
    return r


def reapply_classification(s: Session, rule: "Classification") -> int:
    """Re-apply a classification rule to the receiving base (Asset table).

    Finds every Asset whose description matches the rule's pattern (respecting
    the optional company filter) and recomputes its category/model using the
    full rule set (longest matching pattern wins). Returns how many assets were
    updated. Called whenever a model rule is created/edited so the base stays in
    sync with the registered models.
    """
    pattern = (rule.description_pattern or "").strip()
    if not pattern:
        return 0

    q = select(Asset).where(func.upper(Asset.description).like(f"%{pattern.upper()}%"))
    if rule.company:
        q = q.where(func.upper(Asset.company) == rule.company.upper())

    updated = 0
    for a in s.scalars(q).all():
        cat, model, _company = classify(s, a.description or "", a.company or "")
        if cat == "NÃO CLASSIFICADA":
            continue
        if a.category != cat or a.model != model:
            a.category = cat
            a.model = model
            updated += 1
    s.flush()
    return updated


# ── Asset helpers ─────────────────────────────────────────────────

def find_asset(s: Session, r: dict) -> Asset | None:
    """Find an existing Asset by any identifier field."""
    ors = []
    for field, key in (
        (Asset.asset_id, "asset_id"),
        (Asset.asset_number, "ativo"),
        (Asset.tag_number, "etiqueta"),
        (Asset.serial_number, "numero_serie"),
    ):
        v = (r.get(key) or "").strip()
        if v:
            ors.append(func.upper(field) == v.upper())
    return s.scalar(select(Asset).where(or_(*ors)).limit(1)) if ors else None


def upsert_asset(s: Session, r: dict) -> Asset:
    """Find-or-create an Asset and update its fields from a result dict."""
    a = find_asset(s, r)
    if not a:
        a = Asset()
        s.add(a)
    for attr, key in [
        ("company", "empresa"),
        ("book_type_code", "book_type_code"),
        ("asset_id", "asset_id"),
        ("asset_number", "ativo"),
        ("tag_number", "etiqueta"),
        ("serial_number", "numero_serie"),
        ("description", "descricao"),
        ("category", "categoria"),
        ("model", "modelo"),
        ("source", "fonte"),
    ]:
        if r.get(key) not in (None, ""):
            setattr(a, attr, str(r.get(key)))
    try:
        if r.get("custo_asset") not in (None, ""):
            a.cost = Decimal(str(r["custo_asset"]))
    except (ValueError, TypeError, ArithmeticError):
        pass
    try:
        if r.get("dpis"):
            a.dpis = date.fromisoformat(r["dpis"])
    except (ValueError, TypeError):
        pass
    s.flush()
    return a


def asset_dict(a: Asset) -> dict[str, Any]:
    """Serialize an Asset ORM instance to a plain dict."""
    return {
        "asset_db_id": a.id,
        "empresa": a.company,
        "book_type_code": a.book_type_code,
        "imobilizado": a.asset_id or a.asset_number,
        "ativo": a.asset_number,
        "etiqueta": a.tag_number,
        "numero_serie": a.serial_number,
        "descricao": a.description,
        "categoria": a.category,
        "modelo": a.model,
        "custo_asset": float(a.cost) if a.cost is not None else None,
        "dpis": a.dpis.isoformat() if a.dpis else None,
        "fonte": a.source,
    }


def cycle_dict(c: ReceiptCycle) -> dict[str, Any]:
    """Serialize a ReceiptCycle (with its related Asset) to a plain dict."""
    d = {
        "id": c.id,
        "cycle_number": c.cycle_number,
        "data_recebimento": c.received_date.isoformat(),
        "semana": c.iso_week,
        "status": c.status,
        "local": c.location.name if c.location else "",
        "local_id": c.location_id,
        "lote": c.lot_number,
        "aberto": c.open,
    }
    d.update(asset_dict(c.asset))
    return d


def local_search_one(s: Session, q: str) -> dict:
    """Search the local asset base for a single identifier."""
    q = q.strip()
    row = s.scalar(
        select(LocalAsset)
        .where(
            LocalAsset.active == True,  # noqa: E712
            or_(
                func.upper(LocalAsset.asset_number) == q.upper(),
                func.upper(LocalAsset.tag_number) == q.upper(),
                func.upper(LocalAsset.serial_number) == q.upper(),
            ),
        )
        .limit(1)
    )
    if not row:
        return {
            "pesquisado": q,
            "encontrado": False,
            "erro": "Não encontrado",
            "fonte": "BASE LOCAL",
        }
    return {
        "pesquisado": q,
        "encontrado": True,
        "empresa": row.company,
        "ativo": row.asset_number,
        "imobilizado": row.asset_number,
        "asset_id": row.asset_number,
        "etiqueta": row.tag_number,
        "numero_serie": row.serial_number,
        "descricao": row.description,
        "dpis": row.acquisition_date.isoformat() if row.acquisition_date else None,
        "fonte": "BASE LOCAL",
    }


# ── XLSX export ───────────────────────────────────────────────────

_XLSX_COLS = [
    "pesquisado", "encontrado", "empresa", "book_type_code", "categoria",
    "modelo", "descricao", "ativo", "asset_id", "etiqueta", "numero_serie",
    "custo_asset", "dpis", "local_atribuido", "conta_despesas", "fonte", "erro",
]


def xlsx_response(rows: list[dict], name: str) -> StreamingResponse:
    """Build an XLSX file from a list of dicts and return as streaming response."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    # Header row
    ws.append([x.replace("_", " ").title() for x in _XLSX_COLS])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="AB4807")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for r in rows:
        row_data = []
        for c in _XLSX_COLS:
            val = r.get(c, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            row_data.append(val)
        ws.append(row_data)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto-width
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(50, max(12, width))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )
