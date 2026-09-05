"""Public assets router — EBS lookup without user auth (uses service credentials)."""
from __future__ import annotations

import io
import os
import threading
import time

from fastapi import APIRouter, Request, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel, field_validator
from sqlalchemy import text

from config import get_settings
from db.portal import SessionLocal
from integracoes.ebs_service import login as ebs_login, search_many as ebs_search_many

_cfg = get_settings()
router = APIRouter(prefix="/api/public-assets", tags=["Consulta pública EBS"])

_auth_lock = threading.Lock()
_auth_cache: dict = {"value": None, "expires": 0.0}


# ── Pydantic models ───────────────────────────────────────────────

class PublicQueryIn(BaseModel):
    identificadores: list[str]

    @field_validator("identificadores")
    @classmethod
    def clean_ids(cls, v: list[str]) -> list[str]:
        seen: set[str] = set()
        out: list[str] = []
        for raw in v:
            item = str(raw).strip()
            if item and item not in seen:
                seen.add(item)
                out.append(item)
        if not out:
            raise ValueError("Informe ao menos um identificador.")
        return out


# ── Internal helpers ──────────────────────────────────────────────

def _credential(name: str) -> str:
    directory = _cfg.CREDENTIALS_DIRECTORY
    path = os.path.join(directory, name) if directory else ""
    if not path or not os.path.isfile(path):
        raise RuntimeError(f"Credencial protegida não carregada: {name}")
    with open(path, "r", encoding="utf-8") as handle:
        value = handle.read().strip()
    if not value:
        raise RuntimeError(f"Credencial protegida vazia: {name}")
    return value


def _auth(force: bool = False):
    now = time.time()
    with _auth_lock:
        if not force and _auth_cache["value"] and _auth_cache["expires"] > now:
            return _auth_cache["value"]
        value = ebs_login(
            _credential("ebs_public_username"),
            _credential("ebs_public_password"),
        )
        _auth_cache["value"] = value
        _auth_cache["expires"] = now + 600
        return value


def _ip(req: Request) -> str:
    forwarded = req.headers.get("x-forwarded-for", "")
    return (forwarded.split(",")[0].strip() if forwarded else (req.client.host if req.client else ""))[:80]


def _normal(value) -> str:
    return " ".join(str(value or "").strip().upper().split())


def _company(value) -> str:
    value = _normal(value)
    for name in ("RENNER", "YOUCOM", "CAMICADO"):
        if name in value:
            return name
    return value.replace("FA_", "")


def _rules(db):
    """Load classification rules as a lookup dict."""
    rows = db.execute(
        text(
            "SELECT description_pattern, company, category, model "
            "FROM classifications WHERE active=true ORDER BY id DESC"
        )
    ).mappings().all()
    result = {}
    for row in rows:
        key = (_normal(row["description_pattern"]), _normal(row["company"]))
        result.setdefault(key, dict(row))
    return result


def _apply_rule(row: dict, rule_map: dict) -> dict:
    desc = _normal(row.get("descricao"))
    comp = _company(row.get("empresa") or row.get("book_type_code"))
    rule = rule_map.get((desc, _normal(comp))) or rule_map.get((desc, ""))
    row["empresa"] = comp
    row["categoria"] = rule["category"] if rule else "NÃO CLASSIFICADA"
    row["modelo"] = rule["model"] if rule else (row.get("descricao") or "")
    return row


def _snow(row: dict) -> dict:
    dpis = row.get("dpis") or ""
    return {
        "serial_number": row.get("numero_serie") or "",
        "model": row.get("modelo") or "",
        "asset_tag": row.get("etiqueta") or "",
        "model_category": row.get("categoria") or "",
        "stockroom": "SPARE - CD324",
        "state": "In stock",
        "substate": "Available",
        "acquisition_method": "Purchase",
        "aisle_and_space": "",
        "company": row.get("empresa") or "",
        "cost": row.get("custo_asset") if row.get("custo_asset") is not None else "",
        "expenditure_type": "Capex",
        "purchased": dpis,
        "quantity": 1,
        "depreciation": "SL 5 Years",
        "depreciation_effective_date": dpis,
        "descricao_ebs": row.get("descricao") or "",
        "ativo": row.get("ativo") or row.get("asset_id") or "",
    }


def _audit(ip: str, ids: list, found: int, missing: int, exported: bool, outcome: str = "SUCCESS", error: str = ""):
    try:
        with SessionLocal.begin() as db:
            db.execute(
                text(
                    "INSERT INTO public_ebs_query_audit "
                    "(ip_address, identifiers_count, found_count, missing_count, "
                    "exported, outcome, error_message, created_at) "
                    "VALUES (:ip, :total, :found, :missing, :exported, :outcome, :error, now())"
                ),
                {
                    "ip": ip,
                    "total": len(ids),
                    "found": found,
                    "missing": missing,
                    "exported": exported,
                    "outcome": outcome,
                    "error": str(error)[:500],
                },
            )
    except Exception:
        pass


def _execute(ids: list[str], req: Request, exported: bool = False) -> list[dict]:
    ip = _ip(req)
    try:
        raw = ebs_search_many(_auth(), ids)
        if raw and all(
            "Sessão EBS expirada" in str(x.get("erro", ""))
            for x in raw if isinstance(x, dict)
        ):
            raw = ebs_search_many(_auth(True), ids)

        with SessionLocal() as db:
            rule_map = _rules(db)
            converted = [_apply_rule(dict(row), rule_map) for row in raw]

        found = sum(1 for row in converted if row.get("encontrado"))
        _audit(ip, ids, found, len(ids) - found, exported)
        return converted

    except Exception as exc:
        _audit(ip, ids, 0, len(ids), exported, "ERROR", exc)
        print(f"[PUBLIC_EBS_INTEGRATION_ERROR] {type(exc).__name__}: {exc}", flush=True)
        raise HTTPException(
            502,
            "Integração EBS temporariamente indisponível. Tente novamente mais tarde.",
        )


# ── Endpoints ─────────────────────────────────────────────────────

@router.get("/health")
def health():
    try:
        user = _credential("ebs_public_username")
        return {
            "ok": True,
            "module": "public-assets",
            "authentication_required": False,
            "source": "EBS",
            "uses_ebs": True,
            "stored_as_systemd_encrypted_credential": True,
            "credential_configured": bool(user),
            "read_only": True,
            "rate_limit": False,
        }
    except Exception as exc:
        return {
            "ok": False,
            "module": "public-assets",
            "source": "EBS",
            "credential_configured": False,
            "error": str(exc),
        }


@router.post("/convert")
def convert(body: PublicQueryIn, req: Request):
    rows = _execute(body.identificadores, req)
    output = []
    for row in rows:
        entry = {
            "pesquisado": row.get("pesquisado", ""),
            "encontrado": bool(row.get("encontrado")),
        }
        if row.get("encontrado"):
            entry.update(_snow(row))
        else:
            entry["erro"] = row.get("erro", "Não encontrado")
        output.append(entry)

    found = sum(1 for row in output if row["encontrado"])
    return {
        "resultados": output,
        "total": len(body.identificadores),
        "encontrados": found,
        "nao_encontrados": len(body.identificadores) - found,
        "origem": "EBS",
        "autenticacao": "CREDENCIAL PROTEGIDA DO SERVIÇO",
    }


@router.post("/export")
def export(body: PublicQueryIn, req: Request):
    rows = [
        _snow(row)
        for row in _execute(body.identificadores, req, True)
        if row.get("encontrado")
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "ServiceNow"

    cols = [
        ("serial_number", "Serial Number"),
        ("model", "Model"),
        ("asset_tag", "Asset tag"),
        ("model_category", "Model category"),
        ("stockroom", "Stockroom"),
        ("state", "State"),
        ("substate", "Substate"),
        ("acquisition_method", "Acquisition method"),
        ("aisle_and_space", "Aisle and space"),
        ("company", "Company"),
        ("cost", "Cost"),
        ("expenditure_type", "Expenditure type"),
        ("purchased", "Purchased"),
        ("quantity", "Quantity"),
        ("depreciation", "Depreciation"),
        ("depreciation_effective_date", "Depreciation effective date"),
    ]

    ws.append([label for _, label in cols])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="AB4807")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(key, "") for key, _ in cols])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(45, max(12, width))

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ativos_ebs_servicenow.xlsx"'},
    )
